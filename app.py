import io
import json
import sqlite3
import hashlib
import uuid
import random
import smtplib
import time
import re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from PIL import Image

# ------------------------------------------------------------------------------
# 🔑 API Key 및 이메일 연동 (Secrets에서 불러오기)
# ------------------------------------------------------------------------------
GEMINI_API_KEY = st.secrets.get("GEMINI_API_KEY", "")
SMTP_EMAIL = st.secrets.get("SMTP_EMAIL", "")       # 발송용 Gmail 주소
SMTP_PASSWORD = st.secrets.get("SMTP_PASSWORD", "") # Gmail 앱 비밀번호

MARGIN_RATE = 0.10          # 자재 마진율 10%
LABOR_RATE_MAIN = 260000    # 메인 차단기 노무비
LABOR_RATE_BRANCH = 15000   # 분기 차단기 노무비
EXTRA_SHIPPING = 100000     # 현장 운반비 및 양중비

DEFAULT_COLUMNS = ["분전반명", "구분", "종류", "극수", "용량", "부하명", "수량", "단가"]

# 🚫 대폭 확장된 금지어 및 비속어 목록 (무의미한 단어, 초성 욕설, 변형어, 테스트 문구 전면 차단)
FORBIDDEN_WORDS = [
    "아직없음", "없음", "테스트", "모름", "미정", "개인", "임시", "무명", "blank",
    "test", "none", "null", "admin", "undefined", "무소속", "아무나", "아무개", "익명",
    "aaa", "bbb", "ccc", "asdf", "qwer", "1234", "0000",
    "가나다라", "나다라마", "마바사아", "자차카타", "파하가나",
    "시발", "씨발", "씨팔", "씨뱔", "시뱔", "쌰발", "씨바", "시바", "씨빨", "시빨",
    "ㅅㅂ", "ㅆㅂ", "ㅅㅣ발", "씨1발", "시~발", "시.발", "씨.발",
    "병신", "뼝신", "ㅂㅅ", "ㅂ~ㅅ", "ㅄ",
    "개새끼", "개세끼", "개새키", "개쌔끼", "개쉐끼", "ㄱ새끼", "ㄱ새키", "ㄱㅅㄲ",
    "존나", "존냣", "졸라", "쥰나", "ㅈㄴ", "ㅈㄹ", "지랄", "지랠", "짏랄",
    "좆", "좃", "조까", "젓까", "좆까", "좆같", "좃같", "ㅈ까",
    "새끼", "쌔끼", "새키", "떽", "등신", "미친", "미츼", "미칲", "바보", "멍청",
    "년", "놈", "창녀", "창놈", "걸레", "육시랄", "염병", "십팔", "십8", "18년", "18놈",
    "느금마", "느엠창", "애비", "애미", "패륜", "엠창"
]

# ------------------------------------------------------------------------------
# 🌐 접속자 IP 주소 추출 함수
# ------------------------------------------------------------------------------
def get_remote_ip():
    try:
        headers = st.context.headers
        if "X-Forwarded-For" in headers:
            return headers["X-Forwarded-For"].split(",")[0].strip()
        elif "x-forwarded-for" in headers:
            return headers["x-forwarded-for"].split(",")[0].strip()
        return "알수없음"
    except Exception:
        return "127.0.0.1"

# ------------------------------------------------------------------------------
# 📧 이메일 인증번호 발송 함수
# ------------------------------------------------------------------------------
def send_verification_email(to_email, code):
    if not SMTP_EMAIL or not SMTP_PASSWORD:
        return False, "서버의 이메일 발송 설정(Secrets)이 완료되지 않았습니다."
    
    try:
        msg = MIMEMultipart()
        msg['From'] = f"대한일렉트릭 <{SMTP_EMAIL}>"
        msg['To'] = to_email
        msg['Subject'] = "[대한일렉트릭] 회원가입 이메일 인증번호 안내"
        
        body = f"""
        안녕하세요, 대한일렉트릭 견적 프로그램입니다.
        
        회원가입 신청을 위한 이메일 인증번호는 다음과 같습니다.
        
        ■ 인증번호: {code}
        
        프로그램 회원가입 창에 위 6자리 번호를 정확히 입력해 주세요.
        감사합니다.
        """
        msg.attach(MIMEText(body, 'plain'))
        
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(SMTP_EMAIL, SMTP_PASSWORD)
        server.send_message(msg)
        server.quit()
        return True, "인증번호가 이메일로 성공적으로 발송되었습니다!"
    except Exception as e:
        return False, f"이메일 발송 실패: {e}"

# ------------------------------------------------------------------------------
# 🗄️ Database (SQLite) 설정 및 초기화
# ------------------------------------------------------------------------------
DB_FILE = "users.db"

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    c.execute('''
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            password TEXT NOT NULL,
            name TEXT NOT NULL,
            email TEXT,
            phone TEXT,
            role TEXT NOT NULL,
            status TEXT NOT NULL,
            ip_address TEXT,
            session_id TEXT,
            pin_code TEXT,
            expires_at TEXT,
            created_at TEXT NOT NULL
        )
    ''')
    
    for col in ["email", "phone", "ip_address", "session_id", "pin_code", "expires_at"]:
        try:
            c.execute(f"ALTER TABLE users ADD COLUMN {col} TEXT")
        except Exception:
            pass

    c.execute('''
        CREATE TABLE IF NOT EXISTS usage_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL,
            file_name TEXT NOT NULL,
            item_count INTEGER NOT NULL,
            ip_address TEXT,
            analyzed_at TEXT NOT NULL
        )
    ''')
    
    try:
        c.execute("ALTER TABLE usage_logs ADD COLUMN ip_address TEXT")
    except Exception:
        pass
    
    admin_id = "syd1007"
    admin_pass = hashlib.sha256("kmj851007".encode()).hexdigest()
    
    c.execute("SELECT * FROM users WHERE username=?", (admin_id,))
    if not c.fetchone():
        c.execute('''
            INSERT INTO users (username, password, name, email, phone, role, status, ip_address, session_id, pin_code, expires_at, created_at)
            VALUES (?, ?, '최고관리자', 'admin@daehan.com', '010-0000-0000', 'admin', 'approved', '관리자PC', '', '000000', '2099-12-31 23:59:59', ?)
        ''', (admin_id, admin_pass, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    else:
        c.execute('''
            UPDATE users SET password=?, role='admin', status='approved', expires_at='2099-12-31 23:59:59' WHERE username=?
        ''', (admin_pass, admin_id))
    
    conn.commit()
    conn.close()

def hash_pw(pw):
    return hashlib.sha256(pw.encode()).hexdigest()

init_db()

# ------------------------------------------------------------------------------
# 1. 페이지 설정
# ------------------------------------------------------------------------------
st.set_page_config(
    page_title="대한일렉트릭 견적 프로그램",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="collapsed"
)

user_ip = get_remote_ip()

# ------------------------------------------------------------------------------
# 🚨 중복 접속 강제 종료 팝업 모달 정의
# ------------------------------------------------------------------------------
@st.dialog("🚨 보안 경고: 다른 기기 중복 로그인 감지")
def show_duplicate_login_dialog():
    st.error("### ⚠️ 계정이 다른 기기에서 로그인되었습니다!")
    st.markdown("""
    **다른 PC 또는 모바일 기기에서 동일한 계정으로 로그인되어 현재 접속이 강제 종료되었습니다.**
    
    * **사유**: 1계정 1기기 실시간 보안 정책 적용
    * **조치 사항**: 타인에게 비밀번호가 유출되었을 가능성이 있으니, 로그인 후 비밀번호 변경을 권장합니다.
    """)
    st.divider()
    if st.button("확인 및 로그인 화면으로 이동", type="primary", use_container_width=True):
        st.session_state['logged_in'] = False
        st.session_state['user_info'] = None
        st.session_state['show_dup_modal'] = False
        st.rerun()

# ------------------------------------------------------------------------------
# 2. URL 청정화 및 동시 접속 / 5분 비활동 자동 로그아웃 통제
# ------------------------------------------------------------------------------
if "session" in st.query_params:
    st.query_params.clear()

TIMEOUT_SECONDS = 300

if st.session_state.get('logged_in', False):
    current_time = time.time()
    last_activity = st.session_state.get('last_activity', current_time)
    
    if current_time - last_activity > TIMEOUT_SECONDS:
        st.session_state['logged_in'] = False
        st.session_state['user_info'] = None
        st.session_state['timeout_logout'] = True
        st.rerun()
    else:
        st.session_state['last_activity'] = current_time

    current_user_id = st.session_state['user_info']['username']
    current_session_id = st.session_state['user_info'].get('session', '')
    
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT session_id FROM users WHERE username=?", (current_user_id,))
    db_session = c.fetchone()
    conn.close()
    
    if db_session and db_session[0] != current_session_id:
        st.session_state['show_dup_modal'] = True
    else:
        st.components.v1.html("""
            <script>
                setTimeout(function(){
                    window.parent.postMessage({type: 'streamlit:render'}, '*');
                }, 5000);
            </script>
        """, height=0, width=0)

if st.session_state.get('show_dup_modal', False):
    show_duplicate_login_dialog()
    st.stop()

# ------------------------------------------------------------------------------
# 3. 로그인 / 회원가입 UI
# ------------------------------------------------------------------------------
if not st.session_state.get('logged_in', False):
    st.title("⚡ 대한일렉트릭 견적 프로그램")
    st.subheader("🔒 사용자 인증 및 승인이 되어야 접속 가능")
    st.caption(f"🖥️ 현재 접속 IP: **{user_ip}**")
    
    if st.session_state.get('timeout_logout', False):
        st.warning("⚠️ 보안을 위해 5분간 활동이 없어 자동 로그아웃되었습니다. 다시 로그인해 주세요.")
        st.session_state['timeout_logout'] = False
        
    tab1, tab2 = st.tabs(["🔑 로그인", "📝 회원가입 신청"])
    
    with tab1:
        st.markdown("### 로그인")
        
        with st.form(key="login_form"):
            login_id = st.text_input("아이디 (ID)", key="login_id")
            login_pw = st.text_input("비밀번호", type="password", key="login_pw")
            submit_login = st.form_submit_button("로그인하기", type="primary", use_container_width=True)
            
        if submit_login:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute("SELECT username, name, role, status, expires_at FROM users WHERE username=? AND password=?", 
                      (login_id, hash_pw(login_pw)))
            user = c.fetchone()
            conn.close()
            
            if user:
                username, name, role, status, expires_at = user
                
                if role != "admin" and expires_at:
                    try:
                        exp_date = datetime.strptime(expires_at, "%Y-%m-%d %H:%M:%S")
                        if datetime.now() > exp_date:
                            conn = sqlite3.connect(DB_FILE)
                            c = conn.cursor()
                            c.execute("UPDATE users SET status='expired' WHERE username=?", (username,))
                            conn.commit()
                            conn.close()
                            status = "expired"
                    except Exception:
                        pass

                if status == "approved":
                    new_session = str(uuid.uuid4())
                    conn = sqlite3.connect(DB_FILE)
                    c = conn.cursor()
                    c.execute("UPDATE users SET ip_address=?, session_id=? WHERE username=?", (user_ip, new_session, username))
                    conn.commit()
                    conn.close()
                    
                    st.session_state['logged_in'] = True
                    st.session_state['last_activity'] = time.time()
                    st.session_state['user_info'] = {
                        "username": username, 
                        "name": name, 
                        "role": role, 
                        "ip": user_ip, 
                        "session": new_session,
                        "expires_at": expires_at
                    }
                    st.session_state['show_dup_modal'] = False
                    st.success(f"{name}님, 환영합니다!")
                    st.rerun()
                elif status == "pending":
                    st.warning("⏳ 아직 관리자 승인 대기 중인 계정입니다. 관리자가 가입을 승인해야 이용할 수 있습니다.")
                elif status == "expired":
                    st.error("⌛ 7일 무료 체험(또는 사용 기간)이 만료되었습니다. 기간 연장 및 유료 전환은 관리자에게 문의하세요.")
                else:
                    st.error("🚫 사용이 차단되거나 비활성화된 계정입니다.")
            else:
                st.error("❌ 아이디 또는 비밀번호가 일치하지 않습니다.")
                
    with tab2:
        st.markdown("### 회원가입 신청")
        st.info("💡 개인 이메일 인증번호 확인 후 가입 신청이 가능합니다. 신청 후 관리자의 승인을 받으면 7일 무료 체험이 시작됩니다.")
        
        # 0) 아이디(ID) 실시간 엄격 검증
        reg_id = st.text_input("사용할 아이디 (ID - 4자 이상, 영문 포함 필수)", key="reg_id")
        id_val = reg_id.strip()
        id_check_lower = id_val.lower()
        has_id_letter = bool(re.search(r'[a-zA-Z]', id_val)) if id_val else False
        is_repeat_char = bool(re.search(r'(.)\1\1', id_val)) if id_val else False
        
        is_valid_id = len(id_val) >= 4 and has_id_letter and not id_val.isdigit() and not is_repeat_char and not any(f in id_check_lower for f in FORBIDDEN_WORDS)
        
        if id_val:
            if len(id_val) < 4:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 아이디는 최소 4자 이상이어야 합니다.</span>", unsafe_allow_html=True)
            elif id_val.isdigit():
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 숫자만으로 된 아이디는 사용할 수 없습니다. (영문 포함 필수)</span>", unsafe_allow_html=True)
            elif not has_id_letter:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 아이디에 영문자가 최소 1자 이상 포함되어야 합니다.</span>", unsafe_allow_html=True)
            elif is_repeat_char:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 동일한 문자를 3번 이상 연속으로 반복할 수 없습니다.</span>", unsafe_allow_html=True)
            elif any(f in id_check_lower for f in FORBIDDEN_WORDS):
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 금지어나 비속어가 포함된 아이디는 사용할 수 없습니다.</span>", unsafe_allow_html=True)
            else:
                st.markdown("<span style='color:#2e7d32; font-weight:bold;'>✅ 올바른 아이디 형식입니다.</span>", unsafe_allow_html=True)

        # 1) 이름/회사명 실시간 검증 (연속 패턴 가나다라, abcd, ㄱㄱㄱ 등 전면 차단)
        reg_name = st.text_input("이름 / 회사명 (3글자 이상)", key="reg_name")
        name_val = reg_name.strip()
        
        is_jamo_repeat = bool(re.search(r'([ㄱ-ㅎㅏ-ㅣ])\1\1', name_val)) if name_val else False
        is_char_repeat = bool(re.search(r'(.)\1\1', name_val)) if name_val else False # 가가가, 나나나 반복
        is_hangul_seq = bool(re.search(r'(가나다|나다라|다라마|라마바|마바사|바사아|사아자|아자차|자차카|차카타|카타파|타파하)', name_val)) if name_val else False # 가나다라 패턴
        is_alphabet_seq = bool(re.search(r'(abcd|bcde|cdef|defg|efgh|qwer|asdf|zxcv)', name_val.lower())) if name_val else False
        
        clean_name_check = re.sub(r'[^a-zA-Z0-9가-힣]', '', name_val.lower())
        
        if name_val:
            if len(name_val) < 3:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 최소 3글자 이상 입력해야 합니다.</span>", unsafe_allow_html=True)
            elif is_jamo_repeat or is_char_repeat or is_hangul_seq or is_alphabet_seq or any(forbidden in clean_name_check for forbidden in FORBIDDEN_WORDS):
                st.markdown("""
                <div style="color:#d32f2f; font-weight:bold; font-size:14px; margin-top:4px;">
                    ❌ 올바르지 않은 문구(미정, 가나다라, ㄱㄱㄱ, 동일문자 반복, 비속어)는 사용 불가능합니다. <br>
                    ⚠️ 허위/장난 정보 가입 시 관리자에 의해 서비스 이용이 즉시 차단될 수 있습니다.
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown("<span style='color:#2e7d32; font-weight:bold;'>✅ 올바른 이름/회사명 형식입니다.</span>", unsafe_allow_html=True)
        
        # 2) 휴대폰 번호 실시간 검증
        reg_phone = st.text_input("휴대폰 번호 (11자리, 예: 01012345678 또는 010-1234-5678)", key="reg_phone")
        clean_phone = reg_phone.replace("-", "").strip()
        if reg_phone:
            if not clean_phone.isdigit():
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 숫자만 입력해 주세요.</span>", unsafe_allow_html=True)
            elif len(clean_phone) != 11:
                st.markdown(f"<span style='color:#d32f2f; font-weight:bold;'>❌ 휴대폰 번호는 11자리여야 합니다. (현재 {len(clean_phone)}자리 입력됨)</span>", unsafe_allow_html=True)
            else:
                st.markdown("<span style='color:#2e7d32; font-weight:bold;'>✅ 올바른 휴대폰 번호 형식입니다.</span>", unsafe_allow_html=True)

        # 3) 이메일 주소 실시간 검증
        col_e1, col_e2 = st.columns([3, 1])
        with col_e1:
            reg_email = st.text_input("개인 이메일 주소", key="reg_email")
            if reg_email:
                if "@" not in reg_email or "." not in reg_email:
                    st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 올바른 이메일 형식이 아닙니다.</span>", unsafe_allow_html=True)
        with col_e2:
            st.write("")
            st.write("")
            if st.button("✉️ 인증번호 발송", use_container_width=True):
                if not reg_email or "@" not in reg_email:
                    st.error("올바른 이메일 주소를 입력해 주세요.")
                else:
                    code = str(random.randint(100000, 999999))
                    st.session_state['email_code'] = code
                    st.session_state['code_email_target'] = reg_email
                    st.session_state['email_verified'] = False
                    ok, msg = send_verification_email(reg_email, code)
                    if ok:
                        st.success("📩 입력하신 이메일로 6자리 인증번호가 발송되었습니다. 이메일함을 확인해 주세요!")
                    else:
                        st.error(msg)
                        
        col_c1, col_c2 = st.columns([3, 1])
        with col_c1:
            reg_code = st.text_input("이메일로 수신된 6자리 인증번호", key="reg_code")
        with col_c2:
            st.write("")
            st.write("")
            if st.button("✅ 인증번호 확인", use_container_width=True):
                saved_code = st.session_state.get('email_code', None)
                target_email = st.session_state.get('code_email_target', None)
                
                if not saved_code or not reg_code:
                    st.error("인증번호를 발송받은 후 입력해 주세요.")
                elif reg_code == saved_code and reg_email == target_email:
                    st.session_state['email_verified'] = True
                    st.success("🎉 인증번호가 일치합니다! (인증 완료)")
                else:
                    st.session_state['email_verified'] = False
                    st.error("❌ 인증번호가 일치하지 않습니다. 다시 확인해 주세요.")

        if st.session_state.get('email_verified', False):
            st.caption("✅ 이메일 인증이 완료되었습니다.")

        # 4) 비밀번호 실시간 극강 검증
        reg_pw = st.text_input("비밀번호 (6자 이상, 영문+숫자 필수)", type="password", key="reg_pw")
        
        has_letter = bool(re.search(r'[a-zA-Z]', reg_pw)) if reg_pw else False
        has_digit = bool(re.search(r'\d', reg_pw)) if reg_pw else False
        
        pw_repeat = bool(re.search(r'(.)\1\1', reg_pw)) if reg_pw else False
        is_seq_pattern = bool(re.search(r'(1234|2345|3456|4567|5678|6789|0123|abcd|qwer|asdf)', reg_pw.lower())) if reg_pw else False
        
        contains_id = (id_val.lower() in reg_pw.lower() or id_val in reg_pw) if (id_val and len(id_val)>=3) else False
        
        phone_middle = clean_phone[3:7] if len(clean_phone) == 11 else ""
        phone_last = clean_phone[7:] if len(clean_phone) == 11 else ""
        contains_phone = ((phone_middle in reg_pw) or (phone_last in reg_pw)) if (phone_middle and phone_last) else False
        
        email_prefix = reg_email.split("@")[0] if "@" in reg_email else ""
        contains_email = (email_prefix.lower() in reg_pw.lower()) if (email_prefix and len(email_prefix)>=3) else False
        
        is_birth_pattern = bool(re.search(r'(19[5-9]\d|20[0-2]\d)(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])', reg_pw)) or bool(re.search(r'(\d{2})(0[1-9]|1[0-2])(0[1-9]|[12]\d|3[01])', reg_pw)) if reg_pw else False

        is_valid_pw = (
            len(reg_pw) >= 6 and 
            has_letter and 
            has_digit and 
            not reg_pw.isdigit() and 
            not pw_repeat and 
            not is_seq_pattern and 
            not contains_id and 
            not contains_phone and 
            not contains_email and 
            not is_birth_pattern
        )

        if reg_pw:
            if len(reg_pw) < 6:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 비밀번호가 너무 짧습니다. (최소 6자 이상)</span>", unsafe_allow_html=True)
            elif reg_pw.isdigit():
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 숫자만으로 구성된 비밀번호는 사용할 수 없습니다. (영문 포함 필수)</span>", unsafe_allow_html=True)
            elif not (has_letter and has_digit):
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 영문자와 숫자를 반드시 조합해야 합니다.</span>", unsafe_allow_html=True)
            elif pw_repeat:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 동일한 문자를 3번 이상 연속 사용할 수 없습니다. (예: aaa, 111 금지)</span>", unsafe_allow_html=True)
            elif is_seq_pattern:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 1234, abcd, qwer, asdf 등 쉬운 연속 패턴은 사용할 수 없습니다.</span>", unsafe_allow_html=True)
            elif contains_id:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 아이디(ID)와 동일하거나 유사한 단어/숫자는 포함할 수 없습니다.</span>", unsafe_allow_html=True)
            elif contains_phone:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 휴대폰 번호(중간/뒷자리) 숫자는 비밀번호에 사용할 수 없습니다.</span>", unsafe_allow_html=True)
            elif contains_email:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 이메일 아이디 단어는 비밀번호에 포함할 수 없습니다.</span>", unsafe_allow_html=True)
            elif is_birth_pattern:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 생년월일 날짜 형태(예: 900101, 19950520 등)는 사용할 수 없습니다.</span>", unsafe_allow_html=True)
            else:
                st.markdown("<span style='color:#2e7d32; font-weight:bold;'>✅ 아주 안전하고 유효한 비밀번호입니다.</span>", unsafe_allow_html=True)

        reg_pw_confirm = st.text_input("비밀번호 확인", type="password", key="reg_pw_confirm")
        if reg_pw_confirm:
            if reg_pw != reg_pw_confirm:
                st.markdown("<span style='color:#d32f2f; font-weight:bold;'>❌ 입력하신 두 비밀번호가 서로 일치하지 않습니다.</span>", unsafe_allow_html=True)
            else:
                st.markdown("<span style='color:#2e7d32; font-weight:bold;'>✅ 비밀번호가 정확히 일치합니다.</span>", unsafe_allow_html=True)

        if st.button("가입 신청하기", type="primary", use_container_width=True):
            missing_fields = []
            if not id_val: missing_fields.append("아이디")
            if not name_val: missing_fields.append("이름 / 회사명")
            if not reg_phone.strip(): missing_fields.append("휴대폰 번호")
            if not reg_email.strip(): missing_fields.append("이메일 주소")
            if not reg_pw.strip(): missing_fields.append("비밀번호")
            if not reg_pw_confirm.strip(): missing_fields.append("비밀번호 확인")

            if missing_fields:
                st.error(f"❌ 아래 항목이 입력되지 않았습니다: **{', '.join(missing_fields)}**")
            elif not is_valid_id:
                st.error("❌ 아이디 조건을 확인해 주세요. (4자 이상, 영문 포함 필수, 숫자만 사용 및 반복문자 금지)")
            elif len(name_val) < 3:
                st.error("❌ 이름 / 회사명은 최소 3글자 이상 입력해 주세요.")
            elif is_jamo_repeat or is_char_repeat or is_hangul_seq or is_alphabet_seq or any(forbidden in clean_name_check for forbidden in FORBIDDEN_WORDS):
                st.error("❌ 올바른 이름 또는 회사명을 입력해 주세요. (무의미한 문구, 가나다라, ㄱㄱㄱ, 비속어, '미정/없음' 등은 가입 불가능하며 관리자에 의해 차단될 수 있습니다)")
            elif not clean_phone.isdigit() or len(clean_phone) != 11:
                st.error("❌ 휴대폰 번호는 하이픈(-) 포함 여부와 상관없이 숫자 11자리로 입력해 주세요. (예: 01012345678)")
            elif not is_valid_pw:
                st.error("❌ 비밀번호 보안 규칙을 확인해 주세요. (6자 이상, 영문+숫자 필수, ID/휴대폰/이메일/생년월일/연속문자 사용 불가)")
            elif reg_pw != reg_pw_confirm:
                st.error("❌ 비밀번호와 비밀번호 확인이 일치하지 않습니다.")
            elif not st.session_state.get('email_verified', False):
                st.error("❌ 이메일 [✅ 인증번호 확인] 버튼을 누르고 인증을 완료해 주세요.")
            else:
                formatted_phone = f"{clean_phone[:3]}-{clean_phone[3:7]}-{clean_phone[7:]}"
                
                conn = sqlite3.connect(DB_FILE)
                c = conn.cursor()
                
                c.execute("SELECT * FROM users WHERE username=?", (id_val,))
                if c.fetchone():
                    st.error("❌ 이미 존재하거나 사용 중인 아이디입니다.")
                    conn.close()
                else:
                    c.execute('''
                        INSERT INTO users (username, password, name, email, phone, role, status, ip_address, created_at)
                        VALUES (?, ?, ?, ?, ?, 'user', 'pending', ?, ?)
                    ''', (id_val, hash_pw(reg_pw), name_val, reg_email.strip(), formatted_phone, user_ip, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
                    conn.commit()
                    conn.close()
                    
                    if 'email_code' in st.session_state: del st.session_state['email_code']
                    if 'code_email_target' in st.session_state: del st.session_state['code_email_target']
                    if 'email_verified' in st.session_state: del st.session_state['email_verified']
                    
                    st.success("🎉 이메일 인증 및 가입 신청이 성공적으로 완료되었습니다! 관리자가 승인해 주시면 7일 무료 체험 권한이 부여됩니다.")
    st.stop()

# ------------------------------------------------------------------------------
# 4. 메인 프로그램 화면
# ------------------------------------------------------------------------------
user = st.session_state['user_info']

# ⏳ 남은 이용 시간 카운트다운 배지 계산
remaining_badge = ""
badge_color = "#1E88E5"

if user['role'] == 'admin':
    remaining_badge = "👑 관리자 (무제한 이용)"
    badge_color = "#2E7D32"
else:
    expires_str = user.get('expires_at', '')
    if expires_str:
        try:
            exp_dt = datetime.strptime(expires_str, "%Y-%m-%d %H:%M:%S")
            now_dt = datetime.now()
            diff = exp_dt - now_dt
            
            if diff.total_seconds() > 0:
                days = diff.days
                hours = diff.seconds // 3600
                if days > 3:
                    remaining_badge = f"🟢 이용 가능: D-{days} ({days}일 {hours}시간 남음)"
                    badge_color = "#2E7D32"
                else:
                    remaining_badge = f"⚠️ 만료 임박: D-{days} ({days}일 {hours}시간 남음)"
                    badge_color = "#D32F2F"
            else:
                remaining_badge = "⌛ 이용 기간 만료됨"
                badge_color = "#C62828"
        except Exception:
            remaining_badge = "⏳ 만료일 확인 불가"

col_h1, col_h2 = st.columns([6, 4])
with col_h1:
    st.title("⚡ 대한일렉트릭 견적 프로그램")
    st.caption(f"접속 계정: **{user['name']} ({user['username']})** [{user['role'].upper()}] | 접속 IP: **{user.get('ip', user_ip)}**")
with col_h2:
    st.write("")
    if remaining_badge:
        st.markdown(f"""
            <div style="
                background-color: {badge_color};
                color: white;
                font-size: 20px;
                font-weight: 800;
                padding: 10px 18px;
                border-radius: 8px;
                text-align: center;
                box-shadow: 0 4px 6px rgba(0,0,0,0.1);
                margin-bottom: 8px;
            ">
                {remaining_badge}
            </div>
        """, unsafe_allow_html=True)
        
    if st.button("🚪 로그아웃", type="secondary", use_container_width=True):
        st.session_state['logged_in'] = False
        st.session_state['user_info'] = None
        st.rerun()

# ------------------------------------------------------------------------------
# 👑 관리자 전용 메뉴
# ------------------------------------------------------------------------------
if user['role'] == 'admin':
    with st.expander("👑 [관리자 전용] 회원 승인 및 이용 기간(7일/연장) 관리", expanded=True):
        admin_tab1, admin_tab2 = st.tabs(["👥 회원 승인 & 기간 연장 관리", "📜 이용 이력(로그) 보기"])
        
        with admin_tab1:
            st.subheader("회원가입 승인 및 이용 기간 부여/재발급")
            conn = sqlite3.connect(DB_FILE)
            df_users = pd.read_sql_query("SELECT username AS 아이디, name AS 이름, email AS 이메일, phone AS 연락처, role AS 권한, status AS 상태, expires_at AS 만료일시, ip_address AS 접속IP, created_at AS 가입일시 FROM users", conn)
            
            now_dt = datetime.now()
            def calc_remaining_days(row):
                if row['권한'] == 'admin': return "무제한 (관리자)"
                if not row['만료일시']: return "미설정 (대기중)"
                try:
                    exp_dt = datetime.strptime(row['만료일시'], "%Y-%m-%d %H:%M:%S")
                    diff_days = (exp_dt - now_dt).days
                    if diff_days < 0: return "⌛ 만료됨"
                    return f"🟢 {diff_days}일 남음"
                except Exception:
                    return "-"
            
            df_users['남은기간'] = df_users.apply(calc_remaining_days, axis=1)
            
            cols_order = ['아이디', '이름', '상태', '남은기간', '만료일시', '연락처', '이메일', '가입일시', '접속IP']
            st.dataframe(df_users[[c for c in cols_order if c in df_users.columns]], use_container_width=True)
            
            st.markdown("#### ⚙️ 회원 승인 / 상태 및 이용 기간 선택 설정")
            
            non_admin_users = df_users[df_users['권한'] != 'admin']['아이디'].tolist()
            
            if not non_admin_users:
                st.info("💡 현재 가입된 일반 회원(사용자)이 없습니다.")
            else:
                c1, c2, c3, c4 = st.columns([3, 3, 3, 2])
                with c1:
                    target_user = st.selectbox("대상 회원 선택 (일반 회원)", non_admin_users)
                with c2:
                    new_status = st.selectbox("변경할 상태 선택", [
                        "approved (승인 - 접속허용)", 
                        "pending (대기)", 
                        "expired (만료 처리)", 
                        "rejected (차단)"
                    ])
                with c3:
                    period_option = st.selectbox("부여할 이용 기간 (오늘 기준)", [
                        "7일 (기본 무료체험)",
                        "30일 (1개월 연장)",
                        "90일 (3개월 연장)",
                        "180일 (6개월 연장)",
                        "365일 (1년 구독)",
                        "기존 만료일 유지 (변경 안 함)"
                    ])
                with c4:
                    st.write("")
                    st.write("")
                    if st.button("상태 및 기간 적용", type="primary", use_container_width=True):
                        status_code = new_status.split()[0]
                        days_map = {
                            "7일 (기본 무료체험)": 7,
                            "30일 (1개월 연장)": 30,
                            "90일 (3개월 연장)": 90,
                            "180일 (6개월 연장)": 180,
                            "365일 (1년 구독)": 365
                        }
                        
                        c = conn.cursor()
                        if period_option in days_map and status_code == "approved":
                            new_expires = (datetime.now() + timedelta(days=days_map[period_option])).strftime("%Y-%m-%d %H:%M:%S")
                            c.execute("UPDATE users SET status=?, expires_at=? WHERE username=?", (status_code, new_expires, target_user))
                            st.success(f"🎉 [{target_user}] 회원의 상태가 [승인]으로 설정되었으며, 오늘부터 {days_map[period_option]}일간({new_expires}까지) 이용 권한이 부여되었습니다!")
                        else:
                            c.execute("UPDATE users SET status=? WHERE username=?", (status_code, target_user))
                            st.success(f"[{target_user}] 회원의 상태가 [{status_code}]로 즉시 변경되었습니다!")
                        
                        conn.commit()
                        conn.close()
                        st.rerun()
            conn.close()
            
        with admin_tab2:
            st.subheader("도면 분석 및 견적 이용 기록")
            conn = sqlite3.connect(DB_FILE)
            df_logs = pd.read_sql_query("SELECT id AS 번호, username AS 사용자, file_name AS 도면명, item_count AS 추출수량, ip_address AS 접속IP, analyzed_at AS 분석일시 FROM usage_logs ORDER BY id DESC", conn)
            st.dataframe(df_logs, use_container_width=True)
            conn.close()

# ------------------------------------------------------------------------------
# 5. Gemini AI 도면 분석 함수
# ------------------------------------------------------------------------------
def analyze_drawing_with_gemini(image_pil, api_key, file_name, current_user, ip_addr):
    try:
        import google.generativeai as genai
        genai.configure(api_key=api_key)
        
        prompt = """
        이 분전반 도면 이미지를 정확히 분석하여 도면에 실제로 존재하는 부품 요소들만 추출해 JSON 배열로 반환하세요.

        [추출 및 구분 작성 규칙]
        1. 분전반명: 도면에 기재된 분전반 이름 (예: L-1, TQ-1, 분전반 등)
        2. 구분 항목 규칙:
           - 메인 차단기: "MAIN"
           - 분기 차단기: "분기"
           - 부속 자재(콘센트, 단자대, 계량기 등): "MAIN" 또는 "분기"로 구분하거나, "부속"이라는 단어는 절대로 사용하지 마세요.
        3. 종류: MCCB, ELB, 콘센트, 단자대, 계량기 등
        4. 극수 및 용량: 3P, 2P / 50AF/40AT, N.T/E.T 등
        5. 부하명 및 수량, 단가

        [JSON 반환 포맷 예시]
        [
            {"분전반명": "L-1", "구분": "MAIN", "종류": "MCCB", "극수": "3P", "용량": "50AF/40AT", "부하명": "메인", "수량": 1, "단가": 85000},
            {"분전반명": "L-1", "구분": "분기", "종류": "ELB", "극수": "2P", "용량": "30AF/20AT", "부하명": "L1", "수량": 1, "단가": 12500},
            {"분전반명": "L-1", "구분": "분기", "종류": "콘센트", "극수": "", "용량": "2구", "부하명": "콘센트", "수량": 4, "단가": 5000}
        ]
        
        주의: 마크다운 태그(```json 등) 없이 오직 pure JSON 배열만 반환하세요.
        """
        
        available_models = []
        for m in genai.list_models():
            if 'generateContent' in m.supported_generation_methods:
                m_name = m.name.replace("models/", "")
                available_models.append(m_name)

        available_models.sort(key=lambda x: 0 if 'flash' in x else 1)
        
        if not available_models:
            raise Exception("사용 가능한 Gemini AI 모델이 존재하지 않습니다.")

        response = None
        last_error = None

        for model_name in available_models:
            try:
                model = genai.GenerativeModel(model_name)
                response = model.generate_content([prompt, image_pil])
                if response and response.text:
                    break
            except Exception as ex:
                last_error = ex
                continue

        if not response or not response.text:
            raise last_error if last_error else Exception("도면 분석에 실패하였습니다.")

        text = response.text.strip()
        
        if text.startswith("```json"):
            text = text[7:]
        if text.startswith("```"):
            text = text[3:]
        if text.endswith("```"):
            text = text[:-3]
        text = text.strip()
        
        parsed_data = json.loads(text)
        
        try:
            conn = sqlite3.connect(DB_FILE)
            c = conn.cursor()
            c.execute('''
                INSERT INTO usage_logs (username, file_name, item_count, ip_address, analyzed_at)
                VALUES (?, ?, ?, ?, ?)
            ''', (current_user, file_name, len(parsed_data), ip_addr, datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            conn.commit()
            conn.close()
        except Exception:
            pass
            
        return parsed_data
    except Exception as e:
        st.error(f"[{file_name}] 분석 중 오류 발생: {e}")
        return []

def generate_excel_quote(df_items, margin_rate, labor_main, labor_branch, shipping):
    wb = openpyxl.Workbook()
    
    ws_sum = wb.active
    ws_sum.title = "전체_견적요약"
    ws_sum.views.sheetView[0].showGridLines = True
    
    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    ws_sum["B2"] = "대한일렉트릭 통합 견적서"
    ws_sum["B2"].font = Font(size=18, bold=True, color="1F4E78")
    
    headers = ["번호", "분전반명", "자재비(원)", "인건비(원)", "합계금액(원)", "비고"]
    for idx, h in enumerate(headers, start=2):
        c = ws_sum.cell(row=5, column=idx, value=h)
        c.fill = navy_fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")

    ws_mat = wb.create_sheet(title="전체_차단기내역")
    ws_mat.views.sheetView[0].showGridLines = True
    
    m_headers = ["분전반명", "구분", "종류", "극수", "용량", "부하명", "수량", "단가(원)", "금액(원)"]
    for idx, h in enumerate(m_headers, start=1):
        c = ws_mat.cell(row=1, column=idx, value=h)
        c.fill = navy_fill
        c.font = Font(color="FFFFFF", bold=True)
        c.alignment = Alignment(horizontal="center")
        
    for r_idx, row in df_items.iterrows():
        r = r_idx + 2
        ws_mat.cell(row=r, column=1, value=str(row.get("분전반명", "")))
        ws_mat.cell(row=r, column=2, value=str(row.get("구분", ""))).alignment = Alignment(horizontal="center")
        ws_mat.cell(row=r, column=3, value=str(row.get("종류", ""))).alignment = Alignment(horizontal="center")
        ws_mat.cell(row=r, column=4, value=str(row.get("극수", ""))).alignment = Alignment(horizontal="center")
        ws_mat.cell(row=r, column=5, value=str(row.get("용량", ""))).alignment = Alignment(horizontal="center")
        ws_mat.cell(row=r, column=6, value=str(row.get("부하명", "")))
        
        try:
            qty = int(row.get("수량", 0))
        except Exception:
            qty = 0
        try:
            price = int(row.get("단가", 0))
        except Exception:
            price = 0

        q_cell = ws_mat.cell(row=r, column=7, value=qty)
        q_cell.fill = yellow_fill
        q_cell.alignment = Alignment(horizontal="center")
        
        p_cell = ws_mat.cell(row=r, column=8, value=price)
        p_cell.fill = yellow_fill
        p_cell.number_format = "#,##0"
        
        amt_cell = ws_mat.cell(row=r, column=9, value=f"=G{r}*H{r}")
        amt_cell.number_format = "#,##0"

    valid_df = df_items[df_items["분전반명"].astype(str).str.strip() != ""]
    panels = valid_df["분전반명"].unique() if len(valid_df) > 0 else ["기본분전반"]
    
    for idx, p_name in enumerate(panels, start=1):
        r = idx + 5
        ws_sum.cell(row=r, column=2, value=idx).alignment = Alignment(horizontal="center")
        ws_sum.cell(row=r, column=3, value=p_name)
        ws_sum.cell(row=r, column=4, value=f"=SUMIF('전체_차단기내역'!A2:A100, \"{p_name}\", '전체_차단기내역'!I2:I100)").number_format = "#,##0"
        ws_sum.cell(row=r, column=5, value=f"=COUNTIF('전체_차단기내역'!A2:A100, \"{p_name}\")*{labor_branch}").number_format = "#,##0"
        ws_sum.cell(row=r, column=6, value=f"=D{r}+E{r}").number_format = "#,##0"

    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 5, 12)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ------------------------------------------------------------------------------
# 6. 다중 도면 업로드 및 작업 메인 UI
# ------------------------------------------------------------------------------
uploaded_files = st.file_uploader(
    "🖼️ 결선도 도면 여러 장 마우스로 드래그하여 업로드 (PNG, JPG, 복수 선택 가능)", 
    type=["png", "jpg", "jpeg"],
    accept_multiple_files=True
)
st.caption("📌 **지원 파일 형식**: `PNG`, `JPG`, `JPEG` (이미지 용량 최대 1GB 지원, 폴더에서 여러 장 끌어다 넣기 가능)")

if uploaded_files:
    st.info(f"📂 총 **{len(uploaded_files)}개**의 도면 파일이 선택되었습니다.")
    
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.subheader("🖼️ 선택된 도면 미리보기")
        for idx, file in enumerate(uploaded_files):
            img = Image.open(file)
            st.caption(f"📄 도면 {idx+1}: {file.name}")
            st.image(img, use_container_width=True)
            st.divider()
        
    with col2:
        st.subheader("🔍 전체 도면 통합 해석")
        st.info("💡 선명하고 해상도가 높은 이미지일수록 차단기 및 부품 문자 인식률이 대폭 올라갑니다.")
        st.write("")
        
        if st.button("🚀 전체 도면 한번에 분석 시작", type="primary", use_container_width=True):
            all_results = []
            progress_bar = st.progress(0)
            status_text = st.empty()
            
            blank_row = {"분전반명": "", "구분": "", "종류": "", "극수": "", "용량": "", "부하명": "", "수량": None, "단가": None}
            
            for idx, file in enumerate(uploaded_files):
                status_text.text(f" 분석 중 ({idx+1}/{len(uploaded_files)}): {file.name}")
                img = Image.open(file)
                parsed_list = analyze_drawing_with_gemini(img, GEMINI_API_KEY, file.name, user['username'], user_ip)
                
                if parsed_list:
                    if all_results:
                        all_results.append(blank_row)
                    all_results.extend(parsed_list)
                    
                progress_bar.progress((idx + 1) / len(uploaded_files))
                
            status_text.text("모든 도면 분석 완료!")
            
            if all_results:
                st.session_state['extracted_data'] = pd.DataFrame(all_results)
                st.success(f"🎉 총 {len(uploaded_files)}개 도면의 분석이 성공적으로 완료되었습니다!")
            else:
                st.session_state['extracted_data'] = pd.DataFrame(columns=DEFAULT_COLUMNS)
                st.warning("도면에서 추출된 데이터가 없습니다.")

if 'extracted_data' not in st.session_state:
    st.session_state['extracted_data'] = pd.DataFrame(columns=DEFAULT_COLUMNS)

st.divider()

col_t, col_b = st.columns([8, 2])
with col_t:
    st.subheader("📋 추출된 분전반별 차단기 데이터 (통합)")
with col_b:
    if st.button("🗑️ 표 전체 비우기"):
        st.session_state['extracted_data'] = pd.DataFrame(columns=DEFAULT_COLUMNS)
        st.rerun()

edited_df = st.data_editor(
    st.session_state['extracted_data'],
    num_rows="dynamic",
    use_container_width=True
)

st.divider()

excel_data = generate_excel_quote(edited_df, MARGIN_RATE, LABOR_RATE_MAIN, LABOR_RATE_BRANCH, EXTRA_SHIPPING)

st.download_button(
    label="📥 대한일렉트릭 견적서 엑셀 다운로드(.xlsx)",
    data=excel_data,
    file_name="대한일렉트릭_분전반_견적서.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    type="primary"
)
